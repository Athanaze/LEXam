import argparse
import ast
import asyncio
import csv
import os
import time
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from neo4j import GraphDatabase
from litellm import completion, acompletion, embedding


ISSUE_PROMPT = (
    "Return ONLY a Python list of strings named with the legal issues addressed in the exam question. "
    "No explanation, no markdown.\n\nQuestion:\n{question}\n"
)

DOC_SCORE_PROMPT = (
    "You are scoring how helpful a court decision is for answering the exam question. "
    "Score from 0 to 5 (integer). Return ONLY the integer.\n\n"
    "Question:\n{question}\n\n"
    "Court decision:\n{doc}\n"
)

FINAL_ANSWER_PROMPT = (
    "You are an expert in {course_name} and answer in a structured, exam-style manner. "
    "Use only the provided court decisions as sources. If they are insufficient, say so and answer with best effort. "
    "Respond in the same language as the question.\n\n"
    "Question:\n{question}\n\n"
    "Choices (if any):\n{choices}\n\n"
    "Court decisions:\n{docs}\n\n"
    "Answer:"
)


def cosine_similarity(a: List[float], b: List[float]) -> float:
    dot = 0.0
    na = 0.0
    nb = 0.0
    for x, y in zip(a, b):
        dot += x * y
        na += x * x
        nb += y * y
    if na == 0.0 or nb == 0.0:
        return 0.0
    return dot / ((na ** 0.5) * (nb ** 0.5))


def parse_python_list(text: str) -> List[str]:
    cleaned = text.strip()
    try:
        value = ast.literal_eval(cleaned)
        if isinstance(value, list):
            return [str(v).strip() for v in value if str(v).strip()]
    except Exception:
        pass
    # fallback: split lines or commas
    if "\n" in cleaned:
        items = [i.strip("-• \t") for i in cleaned.splitlines()]
        return [i for i in items if i]
    return [i.strip() for i in cleaned.split(",") if i.strip()]


def parse_choices(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [str(v) for v in value]
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        if text.startswith("[") and text.endswith("]"):
            try:
                parsed = ast.literal_eval(text)
                if isinstance(parsed, list):
                    return [str(v) for v in parsed]
            except Exception:
                pass
        if "\n" in text:
            return [line.strip() for line in text.splitlines() if line.strip()]
        if "||" in text:
            return [p.strip() for p in text.split("||") if p.strip()]
        return [text]
    return [str(value)]


def load_questions(path: str, question_field: str, choices_field: str, course_field: str, id_field: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path)
    sheet = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[1]]
    header_map = {h: i for i, h in enumerate(headers)}

    def cell(row, name):
        idx = header_map.get(name)
        if idx is None:
            return None
        return row[idx].value

    rows = []
    for r in sheet.iter_rows(min_row=2, values_only=False):
        q = cell(r, question_field)
        if q is None:
            continue
        rows.append(
            {
                "id": cell(r, id_field),
                "question": str(q),
                "choices": parse_choices(cell(r, choices_field)),
                "course": str(cell(r, course_field)) if cell(r, course_field) is not None else "law",
            }
        )
    return rows


class LiteLLMChat:
    def __init__(self, model_name: str, generation_args: Optional[Dict[str, Any]] = None):
        self.model_name = model_name
        self.generation_args = generation_args or {"temperature": 0}

    def ask(self, message: str) -> str:
        messages = [{"role": "user", "content": message}]
        return completion(model=self.model_name, messages=messages, **self.generation_args).choices[0].message.content

    async def aask(self, message: str) -> str:
        messages = [{"role": "user", "content": message}]
        resp = await acompletion(model=self.model_name, messages=messages, **self.generation_args)
        return resp.choices[0].message.content


def embed_texts(model: str, texts: List[str]) -> List[List[float]]:
    if not texts:
        return []
    resp = embedding(model=model, input=texts)
    return [item["embedding"] for item in resp.data]


def format_doc(node_props: Dict[str, Any]) -> str:
    props = dict(node_props)
    props.pop("embeddings", None)
    # Avoid giant fields by trimming long strings
    for k, v in list(props.items()):
        if isinstance(v, str) and len(v) > 2000:
            props[k] = v[:2000] + "…"
    return str(props)


def get_fulltext_candidates(tx, index_name: str, query: str, limit: int) -> List[Tuple[int, Dict[str, Any], float]]:
    cypher = (
        "CALL db.index.fulltext.queryNodes($index_name, $query) "
        "YIELD node, score "
        "RETURN id(node) AS id, node AS node, score AS score "
        "LIMIT $limit"
    )
    rows = tx.run(cypher, index_name=index_name, query=query, limit=limit)
    return [(r["id"], dict(r["node"]), float(r["score"])) for r in rows]


def get_vector_candidates(
    tx,
    index_name: str,
    query_emb: List[float],
    k: int,
    limit: int,
) -> List[Tuple[int, Dict[str, Any], float]]:
    cypher = (
        "CALL db.index.vector.queryNodes($index_name, $k, $query) "
        "YIELD node, score "
        "MATCH (node)<-[:HAS_EMBEDDING]-(d:court_decisions) "
        "WITH d, max(score) AS score "
        "RETURN id(d) AS id, d AS node, score AS score "
        "ORDER BY score DESC "
        "LIMIT $limit"
    )
    rows = tx.run(cypher, index_name=index_name, k=k, query=query_emb, limit=limit)
    return [(r["id"], dict(r["node"]), float(r["score"])) for r in rows]


def get_neighbors(
    tx,
    node_id: int,
    query_emb: List[float],
    limit: int,
) -> List[Tuple[int, Dict[str, Any], float]]:
    cypher = (
        "MATCH (c:court_decisions)-[]-(n:court_decisions) "
        "WHERE id(c) = $node_id "
        "MATCH (n)-[:HAS_EMBEDDING]->(e:DecisionEmbedding) "
        "WITH n, max(vector.similarity.cosine(e.embedding, $query)) AS sim "
        "RETURN id(n) AS id, n AS node, sim AS sim "
        "ORDER BY sim DESC "
        "LIMIT $limit"
    )
    rows = tx.run(cypher, node_id=node_id, query=query_emb, limit=limit)
    return [(r["id"], dict(r["node"]), float(r["sim"]) if r["sim"] is not None else 0.0) for r in rows]


def retrieve_top_docs(
    driver,
    issue_text: str,
    issue_emb: List[float],
    fulltext_index: str,
    vector_index: str,
    top_k: int,
    vector_candidates_k: int,
    alpha: float,
    database: Optional[str],
) -> List[Dict[str, Any]]:
    with driver.session(database=database) as session:
        bm25_rows = session.execute_read(get_fulltext_candidates, fulltext_index, issue_text, top_k * 5)
        vector_rows = session.execute_read(
            get_vector_candidates,
            vector_index,
            issue_emb,
            max(vector_candidates_k, top_k),
            top_k * 5,
        )

    bm25_by_id = {doc_id: (props, score) for doc_id, props, score in bm25_rows}
    vec_by_id = {doc_id: (props, score) for doc_id, props, score in vector_rows}

    max_bm25 = max([s for _, _, s in bm25_rows], default=1.0)
    max_vec = max([s for _, _, s in vector_rows], default=1.0)

    all_ids = set(bm25_by_id) | set(vec_by_id)
    scored = []
    for doc_id in all_ids:
        props = vec_by_id.get(doc_id, bm25_by_id.get(doc_id))[0]
        vec_score = vec_by_id.get(doc_id, (None, 0.0))[1]
        bm25_score = bm25_by_id.get(doc_id, (None, 0.0))[1]
        vec_norm = vec_score / max_vec if max_vec else 0.0
        bm25_norm = bm25_score / max_bm25 if max_bm25 else 0.0
        combined = alpha * vec_norm + (1 - alpha) * bm25_norm
        scored.append(
            {
                "id": doc_id,
                "props": props,
                "vector_sim": vec_score,
                "bm25": bm25_score,
                "score": combined,
            }
        )

    scored.sort(key=lambda x: x["score"], reverse=True)
    return scored[:top_k]


async def score_docs(llm: LiteLLMChat, question: str, docs: List[Dict[str, Any]], max_parallel: int) -> List[Dict[str, Any]]:
    sem = asyncio.Semaphore(max_parallel)

    async def score_one(doc: Dict[str, Any]) -> Dict[str, Any]:
        async with sem:
            prompt = DOC_SCORE_PROMPT.format(question=question, doc=format_doc(doc["props"]))
            raw = await llm.aask(prompt)
            try:
                score = int(str(raw).strip())
            except Exception:
                score = 0
            doc["help_score"] = score
            return doc

    return await asyncio.gather(*[score_one(d) for d in docs])


async def explore_neighbors(
    driver,
    llm: LiteLLMChat,
    question: str,
    issue_emb: List[float],
    seed_docs: List[Dict[str, Any]],
    neighbor_limit: int,
    neighbor_parallel: int,
    min_sim: float,
    max_delta: float,
    database: Optional[str],
) -> List[Dict[str, Any]]:
    expanded: List[Dict[str, Any]] = []
    for doc in seed_docs:
        base_sim = doc.get("vector_sim", 0.0)
        with driver.session(database=database) as session:
            neighbors = session.execute_read(get_neighbors, doc["id"], issue_emb, neighbor_limit)
        scored_neighbors = []
        for nid, props, sim in neighbors:
            if sim < min_sim or (base_sim - sim) > max_delta:
                continue
            scored_neighbors.append({"id": nid, "props": props, "vector_sim": sim, "bm25": 0.0, "score": sim})
        scored_neighbors.sort(key=lambda x: x["vector_sim"], reverse=True)
        top_neighbors = scored_neighbors[:neighbor_parallel]
        if not top_neighbors:
            continue
        scored = await score_docs(llm, question, top_neighbors, neighbor_parallel)
        expanded.extend(scored)
    return expanded


def format_choices(choices: List[str]) -> str:
    if not choices:
        return ""
    lines = []
    for i, c in enumerate(choices):
        label = chr(ord('A') + i) if i < 26 else str(i + 1)
        lines.append(f"{label}. {c}")
    return "\n".join(lines)


def write_results(path: str, rows: List[Dict[str, Any]]):
    fieldnames = ["id", "question", "choices", "answer", "used_doc_ids", "used_doc_scores"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input_file", required=True)
    parser.add_argument("--output_file", required=True)
    parser.add_argument("--llm", required=True)
    parser.add_argument("--embedding_model", default="together_ai/google/gemma-3-300m-embedding")
    parser.add_argument("--question_field", default="question")
    parser.add_argument("--choices_field", default="choices")
    parser.add_argument("--course_field", default="course")
    parser.add_argument("--id_field", default="id")
    parser.add_argument("--neo4j_uri", default=os.getenv("NEO4J_URI", "bolt://localhost:7687"))
    parser.add_argument("--neo4j_user", default=os.getenv("NEO4J_USER", "neo4j"))
    parser.add_argument("--neo4j_password", default=os.getenv("NEO4J_PASSWORD", ""))
    parser.add_argument("--neo4j_database", default=os.getenv("NEO4J_DATABASE", "neo4j"))
    parser.add_argument("--fulltext_index", default=os.getenv("NEO4J_FULLTEXT_INDEX", "court_decisions_qjt"))
    parser.add_argument("--vector_index", default=os.getenv("NEO4J_VECTOR_INDEX", "decision_embedding_idx"))
    parser.add_argument("--top_k", type=int, default=10)
    parser.add_argument("--vector_candidates_k", type=int, default=200)
    parser.add_argument("--alpha", type=float, default=0.6)
    parser.add_argument("--doc_score_parallel", type=int, default=8)
    parser.add_argument("--neighbor_limit", type=int, default=40)
    parser.add_argument("--neighbor_parallel", type=int, default=4)
    parser.add_argument("--neighbor_min_sim", type=float, default=0.6)
    parser.add_argument("--neighbor_max_delta", type=float, default=0.15)
    parser.add_argument("--time_budget_s", type=int, default=40)
    args = parser.parse_args()

    llm = LiteLLMChat(args.llm)

    driver = GraphDatabase.driver(args.neo4j_uri, auth=(args.neo4j_user, args.neo4j_password))

    questions = load_questions(args.input_file, args.question_field, args.choices_field, args.course_field, args.id_field)
    results = []

    for item in questions:
        start_time = time.time()
        question = item["question"]
        course_name = item["course"]
        choices = item["choices"]

        issue_prompt = ISSUE_PROMPT.format(question=question)
        issue_text = llm.ask(issue_prompt)
        issues = parse_python_list(issue_text)
        if not issues:
            issues = [question]

        all_docs: List[Dict[str, Any]] = []
        for issue in issues:
            if time.time() - start_time > args.time_budget_s:
                break
            issue_emb = embed_texts(args.embedding_model, [issue])[0]
            top_docs = retrieve_top_docs(
                driver,
                issue,
                issue_emb,
                args.fulltext_index,
                args.vector_index,
                args.top_k,
                args.vector_candidates_k,
                args.alpha,
                args.neo4j_database,
            )
            scored = asyncio.run(score_docs(llm, question, top_docs, args.doc_score_parallel))
            kept = [d for d in scored if d.get("help_score", 0) >= 4]

            if kept and (time.time() - start_time) <= args.time_budget_s:
                expanded = asyncio.run(
                    explore_neighbors(
                        driver,
                        llm,
                        question,
                        issue_emb,
                        kept,
                        args.neighbor_limit,
                        args.neighbor_parallel,
                        args.neighbor_min_sim,
                        args.neighbor_max_delta,
                        args.neo4j_database,
                    )
                )
                kept.extend([d for d in expanded if d.get("help_score", 0) >= 4])

            all_docs.extend(kept)

        # Deduplicate by node id and keep highest help_score
        dedup: Dict[int, Dict[str, Any]] = {}
        for d in all_docs:
            existing = dedup.get(d["id"])
            if not existing or d.get("help_score", 0) > existing.get("help_score", 0):
                dedup[d["id"]] = d

        ranked = sorted(dedup.values(), key=lambda x: x.get("help_score", 0), reverse=True)
        docs_context = "\n\n".join([format_doc(d["props"]) for d in ranked])

        final_prompt = FINAL_ANSWER_PROMPT.format(
            course_name=course_name,
            question=question,
            choices=format_choices(choices),
            docs=docs_context,
        )
        answer = llm.ask(final_prompt)

        results.append(
            {
                "id": item["id"],
                "question": question,
                "choices": "|".join(choices),
                "answer": answer,
                "used_doc_ids": ";".join([str(d["id"]) for d in ranked]),
                "used_doc_scores": ";".join([str(d.get("help_score", 0)) for d in ranked]),
            }
        )

    driver.close()
    write_results(args.output_file, results)


if __name__ == "__main__":
    main()
